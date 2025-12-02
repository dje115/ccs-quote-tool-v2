#!/usr/bin/env python3
"""
Report Export Service
Generates CSV, PDF, and Excel exports for various reports
"""

from io import BytesIO
from typing import Dict, Any, List, Optional
from datetime import datetime
import csv

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ReportExportService:
    """Service for exporting reports in various formats"""
    
    def __init__(self):
        pass
    
    def export_to_csv(self, data: List[Dict[str, Any]], filename: str = "report") -> BytesIO:
        """Export data to CSV format"""
        output = BytesIO()
        
        if not data:
            return output
        
        # Get headers from first row
        headers = list(data[0].keys())
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
        output.seek(0)
        return output
    
    def export_to_pdf(self, title: str, data: List[Dict[str, Any]], 
                     summary: Optional[Dict[str, Any]] = None,
                     filename: str = "report") -> BytesIO:
        """Export data to PDF format using reportlab"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is not installed. Install with: pip install reportlab")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1976d2'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Add title
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Add date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Add summary if provided
        if summary:
            story.append(Paragraph('<b>Summary</b>', styles['Heading2']))
            summary_data = [[k, str(v)] for k, v in summary.items()]
            summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.grey),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Add data table
        if data:
            story.append(Paragraph('<b>Data</b>', styles['Heading2']))
            
            # Get headers
            headers = list(data[0].keys())
            
            # Build table data
            table_data = [headers]
            for row in data:
                table_data.append([str(row.get(h, '')) for h in headers])
            
            # Create table
            data_table = Table(table_data, repeatRows=1)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            story.append(data_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_to_excel(self, data: List[Dict[str, Any]], 
                       summary: Optional[Dict[str, Any]] = None,
                       filename: str = "report") -> BytesIO:
        """Export data to Excel format using openpyxl"""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas and openpyxl are not installed. Install with: pip install pandas openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Header style
        header_fill = PatternFill(start_color="1976d2", end_color="1976d2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        row = 1
        
        # Add summary if provided
        if summary:
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = "Summary"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            for key, value in summary.items():
                ws[f'A{row}'] = str(key)
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
        
        # Add data table
        if data:
            headers = list(data[0].keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            
            # Write data
            for data_row in data:
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row, column=col_idx, value=str(data_row.get(header, '')))
                row += 1
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, start=1):
                max_length = len(header)
                for data_row in data:
                    value = str(data_row.get(header, ''))
                    if len(value) > max_length:
                        max_length = len(value)
                ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output



